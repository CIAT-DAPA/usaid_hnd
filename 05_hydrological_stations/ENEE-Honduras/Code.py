# HERLIN R. ESPINOSA G.
# herlin25@GMAIL.com
# SEP-01-2017

import os
import openpyxl
from openpyxl import Workbook

for base, dirs, files in os.walk('Input'):
    files

    for x in range(0, len(files)):
        InputFile = files[x]

        book = Workbook()
        sheet = book.active
        sheet['A1'] = 'ADDRESS'
        sheet['B1'] = 'GROUP'
        sheet['C1'] = 'CHAN'
        sheet['D1'] = 'BAUD'
        sheet['E1'] = 'SIGNAL'
        sheet['F1'] = 'NOISE'
        sheet['G1'] = 'QUALITY'
        sheet['H1'] = 'FREQ'
        sheet['I1'] = 'CAR TIME'
        sheet['J1'] = 'END TIME'
        sheet['K1'] = 'MSG TIME'
        sheet['L1'] = 'ARM'
        sheet['M1'] = 'SCID'
        sheet['N1'] = 'TYPE'
        sheet['O1'] = 'LEN'
        sheet['P1'] = 'Pcpn15'
        sheet['Q1'] = 'Pcpn30'
        sheet['R1'] = 'Pcpn45'
        sheet['S1'] = 'Pcpn60'
        sheet['T1'] = 'Pdia'
        sheet['U1'] = 'Nvl15'
        sheet['V1'] = 'Nvl30'
        sheet['W1'] = 'Nvl45'
        sheet['X1'] = 'Nvl60'

        doc = openpyxl.load_workbook("Input\\" + InputFile)
        sheet2 = doc.worksheets[0]
        NumRows = sheet2.max_row + 1
        NumColumn = sheet2.max_column + 1
        i = 2
        for x in range(2, NumRows):
            if sheet2.cell(row=x, column=3).value == '11' or sheet2.cell(row=x, column=3).value == '143':
                sheet['A' + str(i)] = sheet2.cell(row=x, column=1).value
                sheet['B' + str(i)] = sheet2.cell(row=x, column=2).value
                sheet['C' + str(i)] = sheet2.cell(row=x, column=3).value
                sheet['D' + str(i)] = sheet2.cell(row=x, column=4).value
                sheet['E' + str(i)] = sheet2.cell(row=x, column=4).value
                sheet['F' + str(i)] = sheet2.cell(row=x, column=6).value
                sheet['G' + str(i)] = sheet2.cell(row=x, column=7).value
                sheet['H' + str(i)] = sheet2.cell(row=x, column=8).value
                sheet['I' + str(i)] = sheet2.cell(row=x, column=9).value
                sheet['J' + str(i)] = sheet2.cell(row=x, column=10).value
                sheet['K' + str(i)] = sheet2.cell(row=x, column=11).value
                sheet['L' + str(i)] = sheet2.cell(row=x, column=12).value
                sheet['M' + str(i)] = sheet2.cell(row=x, column=13).value
                sheet['N' + str(i)] = sheet2.cell(row=x, column=14).value
                sheet['O' + str(i)] = sheet2.cell(row=x, column=15).value

                StrData = sheet2.cell(row=x, column=16).value
                StrData = StrData.strip()
                StrData = StrData.replace('"', '')
                ListData = StrData.split(' ')

                # AQUI OPTENEMOS LA INFO DE Pcpn
                Ind = ListData.index(':Pcpn')
                sheet['P' + str(i)] = ListData[Ind+3] #Pcpn15
                sheet['Q' + str(i)] = ListData[Ind+4] #Pcpn30
                sheet['R' + str(i)] = ListData[Ind+5] #Pcpn45
                sheet['S' + str(i)] = ListData[Ind+6] #Pcpn60

                # AQUI OPTENEMOS LA INFO DE Pdia
                Ind = ListData.index(':Pdia')
                sheet['T' + str(i)] = ListData[Ind+3] #Pdia

                # AQUI OPTENEMOS LA INFO DE Nvl
                Ind = ListData.index(':Nvl')
                sheet['U' + str(i)] = ListData[Ind+3] #Nvl15
                sheet['V' + str(i)] = ListData[Ind+4] #Nvl30
                sheet['W' + str(i)] = ListData[Ind+5] #Nvl45
                sheet['X' + str(i)] = ListData[Ind+6] #Nvl60

                i += 1

        book.save("Output\Output_"+InputFile)